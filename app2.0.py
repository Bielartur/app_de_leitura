import openpyxl as xl
import PySimpleGUI as sg

from datetime import datetime
import os

os.system('cls')

# Ler a planilha
planilha = xl.load_workbook('lista_de_livros.xlsx')
pagina_livros = planilha['Livros']
paginas_por_mes = planilha['Paginas por mes']

# Definir cor de fundo
sg.theme('DarkBlue')

evento = 'Start'

def extrair_dados_planilha():
    livros_em_andamento = []
    lista_paginas_totais = []
    todas_paginas_lidas = []
    informacao_completa_livro = []
    id_livros = []

    for linha in pagina_livros.iter_rows(min_row=2, values_only=True, max_col=7):
        id, _, _, _, livro_em_andamento, paginas_totais, pagina_atual = linha
        if livro_em_andamento:
            paginas_totais = int(paginas_totais)
            if pagina_atual/paginas_totais != 1:
                id_livros.append(id)
                livros_em_andamento.append(livro_em_andamento)
                lista_paginas_totais.append(paginas_totais)
                informacao_completa_livro.append([id, livro_em_andamento, paginas_totais, pagina_atual])               
            todas_paginas_lidas.append(pagina_atual)

    return id_livros, livros_em_andamento, lista_paginas_totais, todas_paginas_lidas, informacao_completa_livro

def extrair_meta():
    i = 0
    for linha in paginas_por_mes.iter_rows(min_row=2, values_only=True, min_col=3, max_row=2):
        _, meta = linha
        if meta:
            i += 1
            meta_anual = meta
    if i == 0:
        paginas_por_mes['D2'] = 3400
        meta_anual = 3400

    planilha.save('lista_de_livros.xlsx')

    return meta_anual

def extrair_mes_ano():
    meses = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    data_atual = datetime.now()
    numero_mes = data_atual.month
    mes_conclusao = meses[numero_mes - 1]
    ano_conclusao = data_atual.year

    return mes_conclusao, ano_conclusao, meses, numero_mes

def extrair_qtd_paginas_lidas_mes():
    paginas_dos_meses = []
    for linha in paginas_por_mes.iter_rows(min_row=2, values_only=True, max_col=3):
        _, paginas_do_mes, _ = linha
        paginas_dos_meses.append(paginas_do_mes)
    return paginas_dos_meses

def extrair_livros_concluidos():
    _, ano_atual, _, _ = extrair_mes_ano()
    anos = [2021, 2022, 2023, 2024]
    if ano_atual not in anos:
        anos.append(ano_atual)

    livros_2020 = []
    livros_2021 = []
    livros_2022 = []
    livros_2023 = []
    livros_2024 = []

    for linha in pagina_livros.iter_rows(min_row=2, values_only=True, max_col=4):
        id, nome_livro, mes, ano = linha
        if ano == 2020:
            livros_2020.append([nome_livro, mes, ano])

        if ano == 2021:
            livros_2021.append([nome_livro, mes, ano])

        if ano == 2022:
            livros_2022.append([nome_livro, mes, ano])

        if ano == 2023:
            livros_2023.append([nome_livro, mes, ano])

        if ano == 2024:
            livros_2024.append([nome_livro, mes, ano])

    todos_livros_concluidos = [livros_2020, livros_2021, livros_2022, livros_2023, livros_2024]

    return todos_livros_concluidos 

def atualizar_pagina(livro_selecionado= str, informacao_completa_livro= list):
    for livro in informacao_completa_livro:
        if livro[1] == livro_selecionado:
            paginas_totais = livro[2]
            pagina_atual = livro[3]
        
    porcentagem_lida = round(pagina_atual/paginas_totais * 100, 2)

    layout_adicionar_pag = [
        [sg.Text(f'Você já leu: {pagina_atual}/{paginas_totais} páginas', font=18), sg.Text(f'Progesso {porcentagem_lida}%', font=18)],
        [sg.Text('Página atual:', font=18), sg.InputText()],
        [sg.Button('Atualizar'), sg.Button('Voltar')]
    ]

    janela_adicionar_pag = sg.Window('Adicionar página', layout=layout_adicionar_pag)

    while True:
        evento, valores = janela_adicionar_pag.Read()
        if evento == sg.WINDOW_CLOSED:
            break

        elif evento == 'Atualizar':
            try:
                nova_pagina_atual = int(valores[0])
                if nova_pagina_atual <= paginas_totais:
                    porcentagem_lida = round(nova_pagina_atual/paginas_totais * 100, 2)

                    if porcentagem_lida == 100:
                        adicionar_aos_concluidos(livro_concluido=livro_selecionado)

                    escrever_novas_pags_na_planilha(livro_para_atualizar=livro_selecionado, nova_pagina_atual=nova_pagina_atual)
                    break
                else:
                    sg.popup('Valor digitado inválido')

            except ValueError as e:
                print(e)
                sg.popup('Valor inválido')

        elif evento == 'Voltar':
            janela_adicionar_pag.close()
            evento = atualizar_livros_em_andamento()
            break

    janela_adicionar_pag.close()

    return evento

def escrever_novas_pags_na_planilha(livro_para_atualizar= str, nova_pagina_atual= int):
    for linha in pagina_livros.iter_rows(min_row=2, values_only=True, max_col=7):
        if linha[4] == livro_para_atualizar:
            id = linha[0] + 1
            pagina_livros[f'G{id}'] = nova_pagina_atual

    planilha.save('lista_de_livros.xlsx')

def adicionar_livro_em_andamento():
    layout_adicionar_livro = [
        [sg.Text('Nome do Livro:'), sg.InputText()],
        [sg.Text('Qtd de páginas:'), sg.InputText()],
        [sg.Button('Adicionar'), sg.Button('Voltar')]
    ]

    janela_adicionar_livro = sg.Window('Adicionar livro', layout=layout_adicionar_livro)

    while True:
        evento, valores = janela_adicionar_livro.Read()
        if evento == sg.WINDOW_CLOSED or evento == 'Voltar':
            break
        elif evento == 'Adicionar':
            nome_livro = valores[0]
            paginas_do_livro = valores[1]
            break

    if evento != sg.WINDOW_CLOSED and evento != 'Voltar':
        for linha in pagina_livros.iter_rows(min_row=2, values_only=True, max_col=5):
            id,_ ,_ ,_ , livro_em_andamento = linha
            if livro_em_andamento:
                id_true = int(id)

        pagina_livros[f'E{id_true + 2}'] = nome_livro
        pagina_livros[f'F{id_true + 2}'] = paginas_do_livro
        pagina_livros[f'G{id_true + 2}'] = 0
        pagina_livros[f'H{id_true + 2}'] = f'=G{id_true + 2}/F{id_true + 2}'

        planilha.save('lista_de_livros.xlsx')

    janela_adicionar_livro.close()

def adicionar_aos_concluidos(livro_concluido= str):
    for linha in pagina_livros.iter_rows(min_row=2, values_only=True, max_col=2):
        id, _ = linha
        if id:
            id_true = int(id)

    mes_conclusao, ano_conclusao, _, _ = extrair_mes_ano()

    pagina_livros[f'A{id_true + 2}'] = id_true + 1
    pagina_livros[f'B{id_true + 2}'] = livro_concluido
    pagina_livros[f'C{id_true + 2}'] = mes_conclusao
    pagina_livros[f'D{id_true + 2}'] = ano_conclusao

    planilha.save('lista_de_livros.xlsx') 

def atualizar_total_paginas(todas_paginas_lidas= int):
    mes_conclusao, _, meses, numero_mes = extrair_mes_ano()
    total_pag_lidas = sum(todas_paginas_lidas)

    for linha in paginas_por_mes.iter_rows(min_row=2, values_only=True, max_col=3):
        mes,_ , total_antigo = linha
        if mes == mes_conclusao:
            paginas_por_mes[f'C{numero_mes + 1}'] = total_pag_lidas
        
        elif mes == meses[numero_mes - 2]:
            paginas_do_mes = total_pag_lidas - total_antigo
            paginas_por_mes[f'B{numero_mes + 1}'] = paginas_do_mes

    planilha.save('lista_de_livros.xlsx')    

def atualizar_livros_em_andamento():
    layout_livros_em_andamento = [
    [sg.Listbox(livros_em_andamento, size=(26, 8), key='-LIST-', enable_events=True, justification='center', font= 18)],
    [sg.Button('Adicionar livro'), sg.Button('Selecionar livro'), sg.Button('Voltar')]
    ]

    janela = sg.Window('Livros desse ano', layout=layout_livros_em_andamento)

    while True:
        evento, valores = janela.Read()
        if evento == sg.WINDOW_CLOSED:
            break
        elif evento == 'Adicionar livro':
            janela.close()
            adicionar_livro_em_andamento()
            break

        elif evento == 'Selecionar livro':
            try:
                livro_selecionado = valores['-LIST-'][0]
                janela.close()
                evento = atualizar_pagina(livro_selecionado=livro_selecionado, informacao_completa_livro=informacao_completa_livro)
                break
            except IndexError:
                sg.popup('É nescessário selecionar algum livro')

        elif evento == 'Voltar':
            janela.close()
            meta = extrair_meta()
            evento = home(meta=meta)
                
    janela.close()

    return evento

def atualizar_meta():
    layout_alterar_meta = [
        [sg.Text('Nova meta:', font=18), sg.InputText(font=18)],
        [sg.Button('Alterar')]
    ]

    janela_alterar_meta = sg.Window('Alterar meta', layout=layout_alterar_meta)

    while True:
        evento, valores = janela_alterar_meta.Read()
        if evento == sg.WINDOW_CLOSED:
            break
        elif evento == 'Alterar':
            try:
                meta = int(valores[0])
                paginas_por_mes['D2'] = meta
                break

            except ValueError:
                sg.popup('Valor inválido')
            
    planilha.save('lista_de_livros.xlsx')
    janela_alterar_meta.close()

    return meta

def mostrar_anos_anteriores(todos_livros_concluidos= list):
    anos = [2020, 2021, 2022, 2023, 2024]
    livros_do_ano = []
    ano_selecionado = 0
    layout_anos_anteriores = [
    [sg.Listbox(anos, size=(20, 6), key='-LIST-', enable_events=True, justification='left', font= 18)],
    [sg.Button('Selecionar ano'), sg.Button('Voltar')]
    ]

    janela_anos_anteriores = sg.Window('Livros concluidos', layout=layout_anos_anteriores)

    while True:
        evento, valores = janela_anos_anteriores.Read()
        if evento == sg.WINDOW_CLOSED:
            break
        elif evento == 'Selecionar ano':
            ano_selecionado = valores['-LIST-'][0]
            if ano_selecionado == 2020:
                livros_2020 = todos_livros_concluidos[0]
                livros_do_ano = livros_2020

                janela_anos_anteriores.close()
                mostrar_livros_concluidos(livros_do_ano=livros_2020, ano=2020)
                break

            elif ano_selecionado == 2021:
                livros_2021 = todos_livros_concluidos[1]
                livros_do_ano = livros_2021

                janela_anos_anteriores.close()
                mostrar_livros_concluidos(livros_do_ano=livros_2021, ano=2021)
                break

            elif ano_selecionado == 2022:
                livros_2022 = todos_livros_concluidos[2]
                livros_do_ano = livros_2022

                janela_anos_anteriores.close()
                mostrar_livros_concluidos(livros_do_ano=livros_2022, ano=2022)
                break

            elif ano_selecionado == 2023:
                livros_2023 = todos_livros_concluidos[3]
                livros_do_ano = livros_2023

                janela_anos_anteriores.close()
                mostrar_livros_concluidos(livros_do_ano=livros_2023, ano=2023)
                break

            elif ano_selecionado == 2024:
                livros_2024 = todos_livros_concluidos[4]
                livros_do_ano = livros_2024

                janela_anos_anteriores.close()
                mostrar_livros_concluidos(livros_do_ano=livros_2024, ano=2024)
                break

        elif evento == 'Voltar':
            janela_anos_anteriores.close()
            meta = extrair_meta()
            evento = home(meta=meta)

    janela_anos_anteriores.close()

    return livros_do_ano, ano_selecionado, evento

def mostrar_livros_concluidos(livros_do_ano = list, ano = int):
    layout_livros_concluidos = [
    [sg.Table(values=livros_do_ano, headings=['Nome', 'Mês', 'Ano'], auto_size_columns=False,
            display_row_numbers=False, col_widths=[20, 18],
            num_rows=20, font=18, justification='center')],
    [sg.Button('Voltar'), sg.Button('Fechar')]
    ]

    janela_livros_concluidos = sg.Window(f'Livros {ano}', layout=layout_livros_concluidos)

    while True:
        evento, valores = janela_livros_concluidos.Read()
        if evento == sg.WINDOW_CLOSED:
            break

        elif evento == 'Voltar':
            janela_livros_concluidos.close()
            todos_livros_concluidos = extrair_livros_concluidos()
            livros_do_ano, ano, _ = mostrar_anos_anteriores(todos_livros_concluidos=todos_livros_concluidos)
            break

        elif evento == 'Fechar':
            break

    janela_livros_concluidos.close()


# Página inicial mostrando os meses e quantas páginas foram lidas por mês
def home(meta= int):
    meses_paginas = [] # Declarando a varialvel meses_paginas
    total_pag_lidas = sum(todas_paginas_lidas) # Faz a soma do total de páginas lidas no ano
    _, _, meses, _ = extrair_mes_ano() # importa uma lista com o nome de todos os meses em inglês
    paginas_dos_meses = extrair_qtd_paginas_lidas_mes() # extrai da planilha a quantidade de página lidas pr mês

    for i, mes in enumerate(meses):
        mes = [mes, paginas_dos_meses[i]]
        if paginas_dos_meses[i] == 0:
            pass
        else:
            meses_paginas.append(mes) # Organizando em uma lista, cada mês com sua respectiva quantidade de páginas lidas

    # Definindo o layout da janela_home
    layout_home = [
    [sg.Text(f'Progresso atual {total_pag_lidas}/{meta}', font=20, colors='white'), sg.Button('Alterar meta')],
    [sg.Table(values=meses_paginas, headings=['Mês', 'Páginas'], auto_size_columns=False,
            display_row_numbers=False, col_widths=[10, 12],
            num_rows=min(15, len(meses_paginas)), font=18, justification='center')],
    [sg.Button('Adicionar livro'), sg.Button('Adicionar páginas'), sg.Button('Livros concluidos')]
    ]

    janela_home = sg.Window('Páginas lidas no ano', layout=layout_home)

    while True:
        evento, valores = janela_home.Read()
        if evento == sg.WINDOW_CLOSED:
            break
        elif evento == 'Adicionar livro':
            janela_home.close()
            adicionar_livro_em_andamento()
            break
        elif evento == 'Adicionar páginas':
            janela_home.close()
            evento = atualizar_livros_em_andamento()
            break
        elif evento == 'Livros concluidos':
            janela_home.close()
            todos_livros_concluidos = extrair_livros_concluidos()
            _, _, evento = mostrar_anos_anteriores(todos_livros_concluidos=todos_livros_concluidos)
            break
        elif evento == 'Alterar meta':
            janela_home.close()
            try:
                meta = atualizar_meta()
            except UnboundLocalError:
                pass
            break
            
    janela_home.close()
    return evento

# Extrair e atualizar informações da planilha 
while evento != sg.WINDOW_CLOSED:
    id, livros_em_andamento, lista_paginas_totais, todas_paginas_lidas, informacao_completa_livro = extrair_dados_planilha()
    atualizar_total_paginas(todas_paginas_lidas=todas_paginas_lidas)
    meta = extrair_meta()
    evento = home(meta=meta)
